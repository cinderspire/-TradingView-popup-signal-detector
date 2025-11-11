#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SPOT-ONLY PAPER TRADING BOT
===========================
Uses EXACT 1,310 combinations from PAPER PAIR PARAMETERS.txt
SPOT TRADING ONLY: No shorts, only BUY long / SELL to close
Can use futures data for analysis but trades SPOT only
$1000 starting balance, $10 order size
===========================
"""

import ccxt
import pandas as pd
import numpy as np
import talib
import time
import json
import os
import signal
import sys
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Any
import warnings
import logging
from dataclasses import dataclass, asdict
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Setup
warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class Position:
    """Active trading position with detailed tracking"""
    pair: str
    strategy: str
    timeframe: Optional[str]
    parameters: Dict[str, Any]
    side: str  # Always 'long' (SPOT only)
    entry_price: float
    quantity: float
    entry_time: datetime
    order_value: float  # $10 per order
    stop_loss: Optional[float] = None
    take_profit: Optional[float] = None
    current_price: float = 0.0
    unrealized_pnl: float = 0.0
    unrealized_pnl_pct: float = 0.0
    open_roi: float = 0.0  # Current ROI for this position
    line_number: int = 0  # Original line in parameters file

@dataclass
class Trade:
    """Completed trade record with detailed tracking"""
    pair: str
    strategy: str
    timeframe: Optional[str]
    parameters: Dict[str, Any]
    side: str  # Always 'long' (SPOT only)
    entry_price: float
    exit_price: float
    quantity: float
    entry_time: datetime
    exit_time: datetime
    order_value: float
    pnl: float
    pnl_pct: float
    roi: float  # Return on Investment for this trade
    exit_reason: str
    line_number: int = 0  # Original line in parameters file

class StrategyExecutor:
    """Execute specific strategy with exact parameters"""
    
    def __init__(self, strategy_name: str, parameters: Dict[str, Any], timeframe: Optional[str] = None):
        self.name = strategy_name
        self.parameters = parameters or {}
        self.timeframe = timeframe or '5m'
    
    def generate_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generate trading signals based on exact strategy configuration"""
        if len(df) < 50:
            return {'action': 'hold', 'confidence': 0.0}
        
        try:
            # Route to specific strategy implementation
            if 'RSI' in self.name:
                return self._rsi_signals(df)
            elif 'MACD' in self.name:
                return self._macd_signals(df)
            elif 'Turtle' in self.name:
                return self._turtle_signals(df)
            elif 'Bollinger' in self.name:
                return self._bollinger_signals(df)
            elif self.name in ['adaptive_grid', 'bollinger_bands', 'breakout_momentum', 
                              'enhanced_momentum', 'mean_reversion', 'multi_timeframe',
                              'pattern_recognition', 'rsi_divergence', 'sentiment_momentum',
                              'support_resistance', 'trend_following', 'turtle_donchian',
                              'volatility_breakout']:
                return self._advanced_strategy_signals(df)
            else:
                return self._generic_signals(df)
                
        except Exception as e:
            logger.error(f"Error generating signals for {self.name}: {e}")
            return {'action': 'hold', 'confidence': 0.0}
    
    def _rsi_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """RSI-based signals with exact parameters"""
        rsi_period = self.parameters.get('rsi_period', 14)
        rsi_oversold = self.parameters.get('rsi_oversold', 30)
        rsi_overbought = self.parameters.get('rsi_overbought', 70)
        
        rsi = talib.RSI(df['close'].values, rsi_period)
        current_rsi = rsi[-1]
        
        if current_rsi < rsi_oversold:
            return {'action': 'buy', 'confidence': 0.8, 'stop_loss': 0.03, 'take_profit': 0.06}
        elif current_rsi > rsi_overbought:
            # SPOT ONLY: No shorts, just hold when overbought
            return {'action': 'hold', 'confidence': 0.7}
        else:
            return {'action': 'hold', 'confidence': 0.3}
    
    def _macd_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """MACD-based signals with exact parameters"""
        fast_period = self.parameters.get('fast_period', self.parameters.get('macd1_fast', 12))
        slow_period = self.parameters.get('slow_period', self.parameters.get('macd1_slow', 26))
        signal_period = self.parameters.get('signal_period', 9)
        
        macd, signal, histogram = talib.MACD(df['close'].values, fast_period, slow_period, signal_period)
        
        if len(macd) < 2:
            return {'action': 'hold', 'confidence': 0.0}
        
        if macd[-1] > signal[-1] and macd[-2] <= signal[-2]:
            return {'action': 'buy', 'confidence': 0.7, 'stop_loss': 0.025, 'take_profit': 0.05}
        elif macd[-1] < signal[-1] and macd[-2] >= signal[-2]:
            # SPOT ONLY: No shorts, just hold on bearish MACD
            return {'action': 'hold', 'confidence': 0.6}
        else:
            return {'action': 'hold', 'confidence': 0.2}
    
    def _turtle_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Turtle trading signals with exact parameters"""
        entry_period = self.parameters.get('entry_period', 20)
        atr_multiplier = self.parameters.get('atr_multiplier', 2.0)
        
        highs = df['high'].rolling(entry_period).max()
        lows = df['low'].rolling(entry_period).min()
        
        current_price = df['close'].iloc[-1]
        
        if current_price > highs.iloc[-2]:  # Breakout above previous high
            return {'action': 'buy', 'confidence': 0.85, 'stop_loss': 0.04, 'take_profit': 0.08}
        elif current_price < lows.iloc[-2]:  # Breakout below previous low
            # SPOT ONLY: No shorts on downward breakouts, just hold
            return {'action': 'hold', 'confidence': 0.7}
        else:
            return {'action': 'hold', 'confidence': 0.1}
    
    def _bollinger_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Bollinger Bands signals with exact parameters"""
        bb_period = self.parameters.get('bb_period', 20)
        bb_std = self.parameters.get('bb_std', 2.0)
        volume_multiplier = self.parameters.get('volume_multiplier', 1.5)
        
        bb_upper, bb_middle, bb_lower = talib.BBANDS(df['close'].values, bb_period, bb_std, bb_std)
        current_price = df['close'].iloc[-1]
        
        # Volume confirmation if specified
        volume_ok = True
        if volume_multiplier > 1:
            volume_ma = df['volume'].rolling(20).mean()
            volume_ok = df['volume'].iloc[-1] > volume_ma.iloc[-1] * volume_multiplier
        
        if current_price < bb_lower[-1] and volume_ok:
            return {'action': 'buy', 'confidence': 0.75, 'stop_loss': 0.025, 'take_profit': 0.05}
        elif current_price > bb_upper[-1] and volume_ok:
            # SPOT ONLY: No shorts on upper band breach, just hold
            return {'action': 'hold', 'confidence': 0.6}
        else:
            return {'action': 'hold', 'confidence': 0.3}
    
    def _advanced_strategy_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Advanced strategy signals using exact parameters"""
        if self.name == 'enhanced_momentum':
            return self._enhanced_momentum_signals(df)
        elif self.name == 'adaptive_grid':
            return self._adaptive_grid_signals(df)
        elif self.name == 'volatility_breakout':
            return self._volatility_breakout_signals(df)
        elif self.name == 'rsi_divergence':
            return self._rsi_divergence_signals(df)
        else:
            return self._generic_advanced_signals(df)
    
    def _enhanced_momentum_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Enhanced momentum with exact parameters"""
        sma_fast = self.parameters.get('sma_fast', 10)
        sma_slow = self.parameters.get('sma_slow', 30)
        rsi_lower = self.parameters.get('rsi_lower', 30)
        rsi_upper = self.parameters.get('rsi_upper', 70)
        volume_multiplier = self.parameters.get('volume_multiplier', 1.2)
        
        sma_fast_values = talib.SMA(df['close'].values, sma_fast)
        sma_slow_values = talib.SMA(df['close'].values, sma_slow)
        rsi = talib.RSI(df['close'].values, 14)
        
        current_price = df['close'].iloc[-1]
        volume_ma = df['volume'].rolling(20).mean()
        volume_ok = df['volume'].iloc[-1] > volume_ma.iloc[-1] * volume_multiplier
        
        if (current_price > sma_fast_values[-1] > sma_slow_values[-1] and 
            rsi[-1] > rsi_lower and rsi[-1] < rsi_upper and volume_ok):
            return {'action': 'buy', 'confidence': 0.7, 'stop_loss': 0.03, 'take_profit': 0.06}
        elif (current_price < sma_fast_values[-1] < sma_slow_values[-1] and 
              rsi[-1] < rsi_upper and rsi[-1] > rsi_lower and volume_ok):
            # SPOT ONLY: No shorts on bearish momentum, just hold
            return {'action': 'hold', 'confidence': 0.6}
        else:
            return {'action': 'hold', 'confidence': 0.2}
    
    def _adaptive_grid_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Adaptive grid strategy"""
        grid_levels = self.parameters.get('grid_levels', 5)
        grid_spacing = self.parameters.get('grid_spacing', 0.02)
        
        # Simple grid logic - buy on dips, sell on rallies
        price_change = df['close'].pct_change().iloc[-1]
        
        if price_change < -grid_spacing:
            return {'action': 'buy', 'confidence': 0.6, 'stop_loss': 0.04, 'take_profit': grid_spacing}
        elif price_change > grid_spacing:
            # SPOT ONLY: No shorts on price rallies, just hold
            return {'action': 'hold', 'confidence': 0.5}
        else:
            return {'action': 'hold', 'confidence': 0.1}
    
    def _volatility_breakout_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Volatility breakout with exact parameters"""
        atr_period = self.parameters.get('atr_period', 14)
        atr_multiplier = self.parameters.get('atr_multiplier', 2.0)
        volume_multiplier = self.parameters.get('volume_multiplier', 1.5)
        
        # Calculate ATR
        high_vals = df['high'].values
        low_vals = df['low'].values
        close_vals = df['close'].values
        
        try:
            atr = talib.ATR(high_vals, low_vals, close_vals, atr_period)
            current_atr = atr[-1]
        except:
            current_atr = df['close'].rolling(atr_period).std().iloc[-1]
        
        # Volume confirmation
        volume_ma = df['volume'].rolling(20).mean()
        volume_ok = df['volume'].iloc[-1] > volume_ma.iloc[-1] * volume_multiplier
        
        price_change = abs(df['close'].iloc[-1] - df['close'].iloc[-2])
        
        if price_change > current_atr * atr_multiplier and volume_ok:
            # Determine direction - SPOT ONLY: Only buy on upward breakouts
            if df['close'].iloc[-1] > df['close'].iloc[-2]:
                return {'action': 'buy', 'confidence': 0.8, 'stop_loss': 0.04, 'take_profit': 0.08}
            else:
                # SPOT ONLY: No shorts on downward volatility breakouts
                return {'action': 'hold', 'confidence': 0.6}
        else:
            return {'action': 'hold', 'confidence': 0.1}
    
    def _rsi_divergence_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """RSI divergence with exact parameters"""
        rsi_period = self.parameters.get('rsi_period', 14)
        lookback_period = self.parameters.get('lookback_period', 20)
        oversold = self.parameters.get('oversold', 30)
        overbought = self.parameters.get('overbought', 70)
        
        rsi = talib.RSI(df['close'].values, rsi_period)
        current_rsi = rsi[-1]
        
        if current_rsi < oversold:
            return {'action': 'buy', 'confidence': 0.7, 'stop_loss': 0.03, 'take_profit': 0.06}
        elif current_rsi > overbought:
            # SPOT ONLY: No shorts on overbought conditions
            return {'action': 'hold', 'confidence': 0.6}
        else:
            return {'action': 'hold', 'confidence': 0.3}
    
    def _generic_advanced_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generic signals for other advanced strategies"""
        # Simple momentum-based approach
        rsi = talib.RSI(df['close'].values, 14)
        current_rsi = rsi[-1]
        
        if current_rsi < 35:
            return {'action': 'buy', 'confidence': 0.6, 'stop_loss': 0.03, 'take_profit': 0.06}
        elif current_rsi > 65:
            # SPOT ONLY: No shorts, just hold when RSI high
            return {'action': 'hold', 'confidence': 0.5}
        else:
            return {'action': 'hold', 'confidence': 0.2}
    
    def _generic_signals(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generic signals for unknown strategies"""
        rsi = talib.RSI(df['close'].values, 14)
        
        if rsi[-1] < 30:
            return {'action': 'buy', 'confidence': 0.6, 'stop_loss': 0.03, 'take_profit': 0.05}
        elif rsi[-1] > 70:
            # SPOT ONLY: No shorts, just hold when overbought
            return {'action': 'hold', 'confidence': 0.4}
        else:
            return {'action': 'hold', 'confidence': 0.1}

class ExactPaperTradingBot:
    """Paper trading bot using EXACT parameters from file"""
    
    def __init__(self):
        self.exchange = ccxt.binance({
            'enableRateLimit': True,
            'timeout': 15000,
            'sandbox': False
        })
        
        # Load exact parameters
        self.load_exact_parameters()
        
        # Trading state
        self.active_positions: Dict[str, Position] = {}
        self.completed_trades: List[Trade] = []
        self.initial_balance = 1000.0  # $1000 starting balance
        self.current_balance = self.initial_balance
        self.order_size = 10.0  # $10 per order
        
        # Data storage
        self.market_data: Dict[str, pd.DataFrame] = {}
        self.last_update: Dict[str, datetime] = {}
        
        # Performance tracking
        self.strategy_performance: Dict[str, Dict] = defaultdict(lambda: {
            'total_trades': 0, 'winning_trades': 0, 'total_pnl': 0.0, 'total_roi': 0.0
        })
        
        # Control flags
        self.running = True
        self.update_interval = 30  # 30 seconds between updates
        self.start_time = datetime.now()
        
        # Setup signal handlers for clean exit
        signal.signal(signal.SIGINT, self.signal_handler)
        signal.signal(signal.SIGTERM, self.signal_handler)
        
        print(f"SPOT-ONLY Paper Trading Bot - Starting Balance: ${self.initial_balance} | Order Size: ${self.order_size}")
        print(f"Loaded {len(self.exact_combinations)} exact strategy combinations from PAPER PAIR PARAMETERS.txt")
        
        # Show strategy breakdown
        strategy_counts = {}
        for combo in self.exact_combinations:
            strategy = combo['strategy']
            strategy_counts[strategy] = strategy_counts.get(strategy, 0) + 1
        
        print(f"Strategy breakdown: {len(strategy_counts)} unique strategies loaded")
        print(f"Top strategies: {sorted(strategy_counts.items(), key=lambda x: x[1], reverse=True)[:5]}")
        print("Starting continuous trading...")
    
    def load_exact_parameters(self):
        """Load exact parameters from parsed file"""
        try:
            with open('exact_paper_parameters.json', 'r') as f:
                self.exact_combinations = json.load(f)
            
            # Silent loading
            
        except Exception as e:
            logger.error(f"Error loading exact parameters: {e}")
            logger.error("Make sure to run parse_exact_parameters.py first!")
            sys.exit(1)
    
    def signal_handler(self, signum, frame):
        """Handle termination signals"""
        self.running = False
        self.save_final_excel_report()
        sys.exit(0)
    
    def fetch_market_data(self, symbol: str, timeframe: str = '5m', limit: int = 100) -> pd.DataFrame:
        """Fetch market data with API limit handling"""
        try:
            # Check cache first
            cache_key = f"{symbol}_{timeframe}"
            now = datetime.now()
            
            if (cache_key in self.last_update and 
                (now - self.last_update[cache_key]).seconds < 30):
                return self.market_data.get(cache_key, pd.DataFrame())
            
            # Fetch OHLCV data
            ohlcv = self.exchange.fetch_ohlcv(symbol, timeframe, limit=limit)
            
            if len(ohlcv) < 10:
                return pd.DataFrame()
            
            df = pd.DataFrame(ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
            df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
            df.set_index('timestamp', inplace=True)
            
            # Cache data
            self.market_data[cache_key] = df
            self.last_update[cache_key] = now
            
            return df
            
        except Exception as e:
            return self.market_data.get(f"{symbol}_{timeframe}", pd.DataFrame())
    
    def calculate_position_quantity(self, price: float) -> float:
        """Calculate quantity based on $10 order size"""
        return self.order_size / price
    
    def open_position(self, combination: Dict, action: str, price: float, 
                     stop_loss: Optional[float] = None, take_profit: Optional[float] = None):
        """Open a new position with exact combination parameters"""
        try:
            if self.current_balance < self.order_size:
                return
            
            quantity = self.calculate_position_quantity(price)
            position_key = f"{combination['pair']}_{combination['strategy']}_{combination['line_number']}"
            
            # Close existing position if any
            if position_key in self.active_positions:
                self.close_position(position_key, price, "Strategy Change")
            
            # SPOT ONLY: Always long positions (buy to open, sell to close)
            if action != 'buy':
                return  # Only process buy signals for SPOT
            side = 'long'
            
            position = Position(
                pair=combination['pair'],
                strategy=combination['strategy'],
                timeframe=combination['timeframe'],
                parameters=combination['parameters'],
                side=side,
                entry_price=price,
                quantity=quantity,
                entry_time=datetime.now(),
                order_value=self.order_size,
                stop_loss=stop_loss,
                take_profit=take_profit,
                current_price=price,
                line_number=combination['line_number']
            )
            
            self.active_positions[position_key] = position
            self.current_balance -= self.order_size
            
            tf_str = f" {combination['timeframe']}" if combination['timeframe'] else ""
            print(f"BUY: {combination['pair']} {combination['strategy']}{tf_str} @ ${price:.6f}")
            
        except Exception as e:
            pass
    
    def close_position(self, position_key: str, exit_price: float, reason: str):
        """Close an existing position"""
        try:
            if position_key not in self.active_positions:
                return
            
            position = self.active_positions[position_key]
            
            # Calculate PnL and ROI (SPOT ONLY: Always long positions)
            pnl = (exit_price - position.entry_price) * position.quantity
            pnl_pct = (exit_price / position.entry_price - 1) * 100
            
            roi = (pnl / position.order_value) * 100
            
            # Create trade record
            trade = Trade(
                pair=position.pair,
                strategy=position.strategy,
                timeframe=position.timeframe,
                parameters=position.parameters,
                side=position.side,
                entry_price=position.entry_price,
                exit_price=exit_price,
                quantity=position.quantity,
                entry_time=position.entry_time,
                exit_time=datetime.now(),
                order_value=position.order_value,
                pnl=pnl,
                pnl_pct=pnl_pct,
                roi=roi,
                exit_reason=reason,
                line_number=position.line_number
            )
            
            self.completed_trades.append(trade)
            self.current_balance += position.order_value + pnl
            
            # Update strategy performance
            perf = self.strategy_performance[position.strategy]
            perf['total_trades'] += 1
            perf['total_pnl'] += pnl
            perf['total_roi'] += roi
            if pnl > 0:
                perf['winning_trades'] += 1
            
            del self.active_positions[position_key]
            
            tf_str = f" {position.timeframe}" if position.timeframe else ""
            print(f"SELL: {position.pair} {position.strategy}{tf_str} PnL: ${pnl:.2f} ROI: {roi:.2f}%")
            
        except Exception as e:
            pass
    
    def update_positions(self):
        """Update all active positions"""
        positions_to_close = []
        
        for position_key, position in self.active_positions.items():
            try:
                df = self.fetch_market_data(position.pair)
                if df.empty:
                    continue
                
                current_price = df['close'].iloc[-1]
                position.current_price = current_price
                
                # Calculate unrealized PnL and ROI (SPOT ONLY: Always long)
                unrealized_pnl = (current_price - position.entry_price) * position.quantity
                unrealized_pnl_pct = (current_price / position.entry_price - 1) * 100
                
                position.unrealized_pnl = unrealized_pnl
                position.unrealized_pnl_pct = unrealized_pnl_pct
                position.open_roi = (unrealized_pnl / position.order_value) * 100
                
                # Check exit conditions (SPOT ONLY: Always long positions)
                exit_reason = None
                
                if position.stop_loss and current_price <= position.stop_loss:
                    exit_reason = "Stop Loss"
                elif position.take_profit and current_price >= position.take_profit:
                    exit_reason = "Take Profit"
                elif (datetime.now() - position.entry_time).days >= 7:
                    exit_reason = "Max Hold Time"
                
                if exit_reason:
                    positions_to_close.append((position_key, current_price, exit_reason))
                    
            except Exception as e:
                pass
        
        for position_key, exit_price, reason in positions_to_close:
            self.close_position(position_key, exit_price, reason)
    
    def scan_for_signals(self):
        """Scan exact combinations for trading signals"""
        signals_processed = 0
        
        # Rotate through all exact combinations in batches
        if not hasattr(self, 'combination_index'):
            self.combination_index = 0
        
        batch_size = 50  # Process 50 combinations per cycle
        start_idx = self.combination_index
        end_idx = min(start_idx + batch_size, len(self.exact_combinations))
        
        if end_idx >= len(self.exact_combinations):
            combinations_batch = (self.exact_combinations[start_idx:] + 
                                self.exact_combinations[:batch_size - (len(self.exact_combinations) - start_idx)])
            self.combination_index = batch_size - (len(self.exact_combinations) - start_idx)
        else:
            combinations_batch = self.exact_combinations[start_idx:end_idx]
            self.combination_index = end_idx
        
        for combination in combinations_batch:
            try:
                # Fetch market data for this pair
                timeframe = combination['timeframe'] or '5m'
                df = self.fetch_market_data(combination['pair'], timeframe)
                
                if df.empty:
                    continue
                
                current_price = df['close'].iloc[-1]
                
                # Create strategy executor with exact parameters
                executor = StrategyExecutor(
                    combination['strategy'], 
                    combination['parameters'], 
                    combination['timeframe']
                )
                
                # Generate signals
                signal = executor.generate_signals(df)
                
                # SPOT ONLY: Only process BUY signals (lower threshold for more signals)
                if signal['action'] == 'buy' and signal['confidence'] > 0.3:
                    position_key = f"{combination['pair']}_{combination['strategy']}_{combination['line_number']}"
                    
                    if position_key not in self.active_positions and self.current_balance >= self.order_size:
                        # Calculate stop loss and take profit prices
                        stop_loss_price = None
                        take_profit_price = None
                        
                        if 'stop_loss' in signal:
                            if signal['action'] == 'buy':
                                stop_loss_price = current_price * (1 - signal['stop_loss'])
                                if 'take_profit' in signal:
                                    take_profit_price = current_price * (1 + signal['take_profit'])
                            else:
                                stop_loss_price = current_price * (1 + signal['stop_loss'])
                                if 'take_profit' in signal:
                                    take_profit_price = current_price * (1 - signal['take_profit'])
                        
                        # Open position
                        self.open_position(combination, signal['action'], current_price, 
                                         stop_loss_price, take_profit_price)
                
                signals_processed += 1
                
            except Exception as e:
                pass
            
            # Rate limiting
            time.sleep(0.02)
        
        # Show processed count for verification
        if signals_processed > 0 and signals_processed % 25 == 0:
            print(f"Processed {signals_processed} combinations this cycle")
    
    def calculate_total_roi(self) -> float:
        """Calculate total ROI from initial balance"""
        return ((self.current_balance - self.initial_balance) / self.initial_balance) * 100
    
    def show_performance(self):
        """Show top 20 strategy performance rankings"""
        if self.strategy_performance:
            print(f"\n=== TOP 20 STRATEGY PERFORMANCE RANKINGS ===")
            sorted_strategies = sorted(self.strategy_performance.items(), key=lambda x: x[1]['total_roi'], reverse=True)
            
            for rank, (strategy, perf) in enumerate(sorted_strategies[:20], 1):
                win_rate = (perf['winning_trades'] / perf['total_trades'] * 100) if perf['total_trades'] > 0 else 0
                avg_roi = (perf['total_roi'] / perf['total_trades']) if perf['total_trades'] > 0 else 0
                total_roi = perf['total_roi']
                
                print(f"#{rank:2d}. {strategy:<25} | {perf['total_trades']:3d} trades | {win_rate:5.1f}% win | {avg_roi:6.2f}% avg | {total_roi:7.2f}% total")
            
            print(f"=" * 85)
            
            # Show current balance and totals
            total_roi = self.calculate_total_roi()
            print(f"Balance: ${self.current_balance:,.2f} | Total ROI: {total_roi:.2f}% | Active: {len(self.active_positions)} | Completed: {len(self.completed_trades)}")
            print(f"=" * 85)
    
    def save_final_excel_report(self):
        """Save comprehensive Excel report with exact parameters"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'spot_only_paper_trading_final_report_{timestamp}.xlsx'
            
            wb = openpyxl.Workbook()
            
            # === OVERVIEW SHEET ===
            ws_overview = wb.active
            ws_overview.title = "Overview"
            
            runtime = datetime.now() - self.start_time
            total_roi = self.calculate_total_roi()
            total_open_roi = sum(pos.open_roi for pos in self.active_positions.values())
            
            overview_data = [
                ["Metric", "Value"],
                ["Runtime", str(runtime).split('.')[0]],
                ["Initial Balance", f"${self.initial_balance:,.2f}"],
                ["Current Balance", f"${self.current_balance:,.2f}"],
                ["Total P&L", f"${self.current_balance - self.initial_balance:,.2f}"],
                ["Total ROI", f"{total_roi:.2f}%"],
                ["Open ROI", f"{total_open_roi:.2f}%"],
                ["Order Size", f"${self.order_size}"],
                ["Active Positions", len(self.active_positions)],
                ["Completed Trades", len(self.completed_trades)],
                ["Exact Combinations", len(self.exact_combinations)],
                ["Source File", "PAPER PAIR PARAMETERS.txt"],
                ["Trading Mode", "SPOT ONLY (No Shorts)"]
            ]
            
            for row_idx, (metric, value) in enumerate(overview_data, 1):
                ws_overview.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
                ws_overview.cell(row=row_idx, column=2, value=value)
            
            # === ACTIVE POSITIONS SHEET ===
            if self.active_positions:
                ws_positions = wb.create_sheet("Active Positions")
                
                pos_headers = ["Pair", "Strategy", "Timeframe", "Side", "Entry Price", "Current Price", 
                              "Order Value", "Unrealized P&L", "Open ROI %", "Entry Time", "Line #", "Parameters"]
                
                for col, header in enumerate(pos_headers, 1):
                    cell = ws_positions.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                
                for row, position in enumerate(self.active_positions.values(), 2):
                    data = [
                        position.pair,
                        position.strategy,
                        position.timeframe or "5m",
                        position.side,
                        f"${position.entry_price:.6f}",
                        f"${position.current_price:.6f}",
                        f"${position.order_value}",
                        f"${position.unrealized_pnl:.2f}",
                        f"{position.open_roi:.2f}%",
                        position.entry_time.strftime('%Y-%m-%d %H:%M:%S'),
                        position.line_number,
                        str(position.parameters)
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws_positions.cell(row=row, column=col, value=value)
                        if col == 8:  # Unrealized P&L
                            if position.unrealized_pnl > 0:
                                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                            elif position.unrealized_pnl < 0:
                                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # === COMPLETED TRADES SHEET ===
            if self.completed_trades:
                ws_trades = wb.create_sheet("Completed Trades")
                
                trade_headers = ["Pair", "Strategy", "Timeframe", "Side", "Entry Price", "Exit Price", 
                               "Order Value", "P&L", "ROI %", "Entry Time", "Exit Time", 
                               "Exit Reason", "Line #", "Parameters"]
                
                for col, header in enumerate(trade_headers, 1):
                    cell = ws_trades.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                
                for row, trade in enumerate(sorted(self.completed_trades, key=lambda x: x.exit_time, reverse=True), 2):
                    data = [
                        trade.pair,
                        trade.strategy,
                        trade.timeframe or "5m",
                        trade.side,
                        f"${trade.entry_price:.6f}",
                        f"${trade.exit_price:.6f}",
                        f"${trade.order_value}",
                        f"${trade.pnl:.2f}",
                        f"{trade.roi:.2f}%",
                        trade.entry_time.strftime('%Y-%m-%d %H:%M:%S'),
                        trade.exit_time.strftime('%Y-%m-%d %H:%M:%S'),
                        trade.exit_reason,
                        trade.line_number,
                        str(trade.parameters)
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws_trades.cell(row=row, column=col, value=value)
                        if col == 8:  # P&L
                            if trade.pnl > 0:
                                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                            elif trade.pnl < 0:
                                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Auto-fit columns
            for ws in wb.worksheets:
                for col in ws.columns:
                    max_length = 0
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            wb.save(filename)
            logger.info(f"Final Excel report saved: {filename}")
            print(f"\n*** EXACT PARAMETERS REPORT SAVED: {filename} ***")
            
        except Exception as e:
            logger.error(f"Error generating final Excel report: {e}")
    
    def run_continuous(self):
        """Run continuous paper trading with exact parameters"""
        # Start trading
        performance_counter = 0
        
        while self.running:
            try:
                # Update existing positions
                self.update_positions()
                
                # Scan exact combinations
                self.scan_for_signals()
                
                # Show performance every 20 iterations (more frequent)
                performance_counter += 1
                if performance_counter % 20 == 0:
                    self.show_performance()
                elif performance_counter % 5 == 0:
                    # Show activity indicator every 5 iterations
                    print(f"Scanning... Balance: ${self.current_balance:.2f} | Active: {len(self.active_positions)}")
                
                # Wait for next update
                time.sleep(self.update_interval)
                
            except KeyboardInterrupt:
                break
            except Exception as e:
                time.sleep(5)
        
        self.save_final_excel_report()

def main():
    """Main function"""
    try:
        bot = ExactPaperTradingBot()
        bot.run_continuous()
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    main()